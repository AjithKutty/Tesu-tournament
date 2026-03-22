"""
Parse a tournament Excel file and output structured JSON data —
one file per division/sheet.

Usage:
    python src/parse_tournament.py                           # default tournament
    python src/parse_tournament.py --tournament path/to/dir  # specific tournament

Reads:  Excel file specified in tournament config
Writes: output/divisions/<SheetName>.json  (one per sheet)
        output/divisions/tournament_index.json  (summary index)
"""

import json
import re
import os
import argparse
import openpyxl
from config import (load_config, get_tournament_name, get_event_names,
                     get_level_categories, get_doubles_events,
                     get_category_order, get_format_overrides)

# ── Format detection constants (not tournament-specific) ─────────

ROUND_NAMES = {
    "Round 1": "Round 1",
    "Round 2": "Round 2",
    "Quarterfinals": "Quarter-Final",
    "Semifinals": "Semi-Final",
    "Final": "Final",
}


def parse_sheet_name(name, event_names, level_categories, doubles_events):
    """Parse sheet name like 'MS C-Main Draw' or 'BS U17-Playoff'."""
    m = re.match(r"^([A-Z]{2})\s+(.+?)-(Main Draw|Playoff)$", name)
    if not m:
        return None
    event_code = m.group(1)
    level = m.group(2)
    draw_type = m.group(3).lower().replace(" ", "_")
    category = level_categories.get(level, "Other")
    full_name = event_names.get(event_code, event_code)
    if level in ("35", "45"):
        full_name += f" {level}+"
    elif level == "V":
        full_name += " Elite"
    else:
        full_name += f" {level}"
    is_doubles = event_code in doubles_events
    return {
        "event_code": event_code,
        "level": level,
        "draw_type": draw_type,
        "category": category,
        "full_name": full_name,
        "code": f"{event_code} {level}",
        "is_doubles": is_doubles,
    }


def sheet_to_filename(sheet_name):
    """Convert sheet name to safe filename: 'MS C-Main Draw' -> 'MS_C-Main_Draw.json'."""
    return sheet_name.replace(" ", "_") + ".json"


def cell_val(cell):
    """Return stripped string value or None."""
    if cell.value is None:
        return None
    s = str(cell.value).strip()
    return s if s else None


def extract_seed(name):
    """Extract seeding from name like 'Player Name [1]' or '[3/4]'."""
    m = re.search(r"\[(\d+(?:/\d+)?)\]", name)
    if m:
        clean_name = re.sub(r"\s*\[\d+(?:/\d+)?\]\s*", "", name).strip()
        return clean_name, m.group(1)
    return name, None


def read_sheet_rows(ws):
    """Read all rows from worksheet, returning list of dicts keyed by column letter."""
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        row_data = {"_row": row[0].row}
        for c in row:
            v = cell_val(c)
            if v is not None:
                row_data[c.column_letter] = v
        rows.append(row_data)
    return rows


# ── Column layout detection ──────────────────────────────────────

def detect_layout(rows):
    """Detect column layout from header rows.

    Returns dict with:
        player_col: column letter for player names (first round / name column)
        club_col: column letter for club names
        status_col: column letter for status (WDN/SUB), or None
        has_flag_col: whether there's a flag column

    Two known layouts:
        2025-style: A=pos, B=status, C=club, D=flag, E=player/rounds
        2026-style: A=pos, B=club, C=player/rounds (no status/flag columns)
    """
    for r in rows:
        # Look for header row with known round names or "Club"
        vals = {col: r.get(col, "") for col in "ABCDEFGHIJKLMNO" if col in r}

        # Find the first column with a round name or numbered round-robin header
        for col in sorted(vals.keys()):
            if col == "A":
                continue
            v = vals[col]
            if v in ROUND_NAMES:
                # Elimination: this column is both the header AND the player name column
                player_col = col
            elif v in ("1", "2", "3"):
                # Round-robin grid: numbered columns are the match grid,
                # player names are in the column before the numbers
                player_col = chr(ord(col) - 1)
            else:
                continue

            # Determine club and status columns relative to player_col
            prev = chr(ord(player_col) - 1)
            prev2 = chr(ord(player_col) - 2)
            if vals.get(prev, "").rstrip() == "Club":
                club_col = prev
                status_col = prev2 if vals.get(prev2, "").rstrip() in ("St.",) else None
                has_flag = False
            elif vals.get(prev2, "").rstrip() == "Club":
                # There's a flag column between club and player
                club_col = prev2
                prev3 = chr(ord(player_col) - 3)
                status_col = prev3 if vals.get(prev3, "").rstrip() in ("St.",) else None
                has_flag = True
            else:
                club_col = prev
                status_col = None
                has_flag = False
            return {
                "player_col": player_col,
                "club_col": club_col,
                "status_col": status_col,
                "has_flag_col": has_flag,
            }

        # Check for "St." marker (2025-style header)
        if vals.get("B") in ("St.", "St. "):
            return {
                "player_col": "E",
                "club_col": "C",
                "status_col": "B",
                "has_flag_col": True,
            }

    # Fallback: try both known layouts
    # Check if column E has player-like data (names with seeds)
    for r in rows:
        if r.get("A", "").isdigit() and r.get("E"):
            return {
                "player_col": "E",
                "club_col": "C",
                "status_col": "B",
                "has_flag_col": True,
            }
        if r.get("A", "").isdigit() and r.get("C"):
            return {
                "player_col": "C",
                "club_col": "B",
                "status_col": None,
                "has_flag_col": False,
            }

    # Default to 2025-style
    return {
        "player_col": "E",
        "club_col": "C",
        "status_col": "B",
        "has_flag_col": True,
    }


# ── Format detection ─────────────────────────────────────────────

def detect_format(rows, layout=None):
    """Detect sheet format: 'elimination', 'round_robin', or 'group_playoff'."""
    if layout is None:
        layout = detect_layout(rows)
    pcol = layout["player_col"]

    for r in rows:
        a_val = r.get("A", "")
        if re.match(r"^[A-Z]{2}\s+.+- Group [A-Z]$", a_val):
            return "group_playoff"
    for r in rows:
        val = r.get(pcol, "")
        if val in ("Round 1", "Quarterfinals", "Semifinals", "Final"):
            return "elimination"
    # Round-robin: check for numbered columns after player col and "Standings" row
    rr_start = chr(ord(pcol) + 1)
    for r in rows:
        if r.get(rr_start) in ("1", "2") and layout["status_col"] and r.get(layout["status_col"]) == "St.":
            return "round_robin"
    # Alternative round-robin detection: "Club" header + numbered columns
    for r in rows:
        if r.get(layout["club_col"]) in ("Club", "Club ") and r.get(rr_start) in ("1", "2"):
            return "round_robin"
    for r in rows:
        if r.get(layout["club_col"]) == "Standings":
            return "round_robin"
    return "elimination"


# ── Player extraction ────────────────────────────────────────────

def parse_doubles_names(name_str, club_str):
    """Parse doubles pair from newline-separated names/clubs."""
    names = [n.strip() for n in name_str.split("\n") if n.strip()]
    clubs = [c.strip() for c in (club_str or "").split("\n") if c.strip()]
    players = []
    for i, n in enumerate(names):
        clean, seed = extract_seed(n)
        players.append({
            "name": clean,
            "club": clubs[i] if i < len(clubs) else None,
            "seed": seed,
        })
    return players


def extract_elimination_players(rows, is_doubles, layout=None):
    """Extract players/pairs from elimination-style draw. Returns list + draw_size."""
    if layout is None:
        layout = detect_layout(rows)
    pcol = layout["player_col"]
    ccol = layout["club_col"]
    scol = layout["status_col"]

    players = []
    data_rows = [r for r in rows if r["_row"] >= 5]
    max_pos = 0

    def _is_bye_name(val):
        """Check if a player name is a bye (e.g., 'Bye', 'Bye 1', 'Bye 5')."""
        return val.lower() == "bye" or val.lower().startswith("bye ")

    if is_doubles:
        prev_row = None
        for r in data_rows:
            a_val = r.get("A")
            p_val = r.get(pcol)
            if a_val and a_val.isdigit():
                pos = int(a_val)
                max_pos = max(max_pos, pos)
                if p_val and not _is_bye_name(p_val):
                    name1_raw = prev_row.get(pcol, "") if prev_row else ""
                    club1 = prev_row.get(ccol, "") if prev_row else ""
                    name2_raw = p_val
                    club2 = r.get(ccol, "")
                    status1 = prev_row.get(scol) if prev_row and scol else None
                    status2 = r.get(scol) if scol else None
                    status = status1 or status2

                    name1, seed1 = extract_seed(name1_raw) if name1_raw else ("", None)
                    name2, seed2 = extract_seed(name2_raw)
                    seed = seed1 or seed2

                    if name1:
                        pair = [
                            {"name": name1, "club": club1},
                            {"name": name2, "club": club2},
                        ]
                        players.append({
                            "position": pos,
                            "players": pair,
                            "seed": seed,
                            "status": status,
                        })
                prev_row = None
            else:
                prev_row = r if r.get(pcol) else None
    else:
        for r in data_rows:
            a_val = r.get("A")
            p_val = r.get(pcol)
            if a_val and a_val.isdigit():
                pos = int(a_val)
                max_pos = max(max_pos, pos)
                if p_val and not _is_bye_name(p_val):
                    name, seed = extract_seed(p_val)
                    players.append({
                        "position": pos,
                        "name": name,
                        "club": r.get(ccol),
                        "seed": seed,
                        "status": r.get(scol) if scol else None,
                    })

    return players, max_pos


def extract_roundrobin_players(rows, is_doubles, layout=None):
    """Extract players from round-robin draw."""
    if layout is None:
        layout = detect_layout(rows)
    pcol = layout["player_col"]
    ccol = layout["club_col"]
    scol = layout["status_col"]

    players = []
    for r in rows:
        if r.get(ccol) == "Standings":
            break
        a_val = r.get("A")
        p_val = r.get(pcol)
        if a_val and a_val.isdigit() and p_val:
            pos = int(a_val)
            if is_doubles and "\n" in p_val:
                pair = parse_doubles_names(p_val, r.get(ccol))
                seed = None
                for p in pair:
                    if p["seed"]:
                        seed = p["seed"]
                        break
                players.append({
                    "position": pos,
                    "players": pair,
                    "seed": seed,
                    "status": r.get(scol) if scol else None,
                })
            else:
                name, seed = extract_seed(p_val)
                if name.lower() != "bye":
                    players.append({
                        "position": pos,
                        "name": name,
                        "club": r.get(ccol),
                        "seed": seed,
                        "status": r.get(scol) if scol else None,
                    })
    return players


def extract_group_playoff(rows, is_doubles, layout=None):
    """Extract groups from group+playoff format."""
    if layout is None:
        layout = detect_layout(rows)
    pcol = layout["player_col"]
    ccol = layout["club_col"]
    scol = layout["status_col"]

    groups = []
    current_group = None
    current_group_rows = []

    for r in rows:
        a_val = r.get("A", "")
        if re.match(r"^[A-Z]{2}\s+.+- Group [A-Z]$", a_val):
            if current_group:
                num_players = len(current_group["players"])
                current_group["_round_schedule"] = extract_roundrobin_schedule(
                    current_group_rows, num_players, layout
                )
                groups.append(current_group)
            group_letter = a_val[-1]
            current_group = {"name": f"Group {group_letter}", "players": []}
            current_group_rows = []
            continue

        if current_group is None:
            continue

        current_group_rows.append(r)

        if r.get(ccol) == "Standings":
            num_players = len(current_group["players"])
            current_group["_round_schedule"] = extract_roundrobin_schedule(
                current_group_rows, num_players, layout
            )
            groups.append(current_group)
            current_group = None
            current_group_rows = []
            continue

        p_val = r.get(pcol)
        if a_val.isdigit() and p_val:
            pos = int(a_val)
            if is_doubles and "\n" in p_val:
                pair = parse_doubles_names(p_val, r.get(ccol))
                seed = None
                for p in pair:
                    if p["seed"]:
                        seed = p["seed"]
                        break
                current_group["players"].append({
                    "position": pos,
                    "players": pair,
                    "seed": seed,
                    "status": r.get(scol) if scol else None,
                })
            else:
                name, seed = extract_seed(p_val)
                if name.lower() != "bye":
                    current_group["players"].append({
                        "position": pos,
                        "name": name,
                        "club": r.get(ccol),
                        "seed": seed,
                        "status": r.get(scol) if scol else None,
                    })

    if current_group:
        num_players = len(current_group["players"])
        current_group["_round_schedule"] = extract_roundrobin_schedule(
            current_group_rows, num_players, layout
        )
        groups.append(current_group)

    return groups


# ── Bracket / match generation ───────────────────────────────────

def get_round_headers(rows, layout=None):
    """Extract ordered round names from header row."""
    if layout is None:
        layout = detect_layout(rows)
    pcol = layout["player_col"]

    for r in rows:
        # Find the header row: it has a known round name in the player column
        val = r.get(pcol, "")
        if val not in ROUND_NAMES:
            continue
        rounds = []
        for col in sorted(r.keys()):
            if col < pcol or col == "_row":
                continue
            v = r[col]
            if v in ROUND_NAMES:
                rounds.append(ROUND_NAMES[v])
            elif v == "Winner":
                continue
        return rounds
    return []


def player_label(player_entry, is_doubles):
    """Get display label for a player/pair entry."""
    if is_doubles:
        names = [p["name"] for p in player_entry.get("players", [])]
        return " / ".join(names)
    return player_entry.get("name", "")


def build_full_bracket(players, draw_size, round_names, is_doubles):
    """
    Build the full bracket structure for all rounds.

    Round 1: actual players from draw positions, paired (1v2, 3v4, ...).
    Later rounds: structural placeholders ('Winner R1-M1 vs Winner R1-M2').
    Bye positions (no player) result in auto-advance for the opponent.
    """
    # Build a position -> player map
    pos_map = {}
    for p in players:
        pos_map[p["position"]] = p

    abbrev_map = {
        "Round 1": "R1",
        "Round 2": "R2",
        "Quarter-Final": "QF",
        "Semi-Final": "SF",
        "Final": "F",
    }

    # Round 1: pair adjacent draw positions
    r1_matches = []
    # Track which matches are empty (both sides Bye) — used to propagate
    # Bye into later rounds so we never have "Bye vs Bye" matches.
    empty_matches = set()  # set of (round_name, match_num)

    for i in range(1, draw_size + 1, 2):
        p1 = pos_map.get(i)
        p2 = pos_map.get(i + 1)

        p1_label = player_label(p1, is_doubles) if p1 else "Bye"
        p2_label = player_label(p2, is_doubles) if p2 else "Bye"

        match_num = len(r1_matches) + 1

        if p1_label == "Bye" and p2_label == "Bye":
            empty_matches.add((round_names[0] if round_names else "Round 1", match_num))
            continue  # Skip double-bye matches entirely

        match = {
            "match": match_num,
            "player1": p1_label,
            "player2": p2_label,
        }

        if p1_label == "Bye" and p2_label != "Bye":
            match["notes"] = f"{p2_label} auto-advances"
        elif p2_label == "Bye" and p1_label != "Bye":
            match["notes"] = f"{p1_label} auto-advances"

        r1_matches.append(match)

    if not round_names:
        return [{"name": "Round 1", "matches": r1_matches}] if r1_matches else []

    rounds = [{"name": round_names[0], "matches": r1_matches}]

    # Build a map of known winners from bye matches so later rounds can
    # use actual player names instead of "Winner X-MN" placeholders.
    # Key: (round_name, match_num) -> winner name
    bye_winners = {}
    for m in r1_matches:
        if "auto-advances" in m.get("notes", ""):
            p1 = m["player1"]
            p2 = m["player2"]
            winner = p1 if p2.startswith("Bye") else p2
            bye_winners[(round_names[0], m["match"])] = winner

    # Later rounds: structural matches, propagating Byes from empty feeders
    prev_round_name = round_names[0]
    prev_match_count = draw_size // 2  # Use expected count, not actual (some were skipped)

    for rnd_idx in range(1, len(round_names)):
        rnd_name = round_names[rnd_idx]
        prev_abbrev = abbrev_map.get(prev_round_name, prev_round_name[:2])
        num_matches = prev_match_count // 2
        if num_matches < 1:
            break

        matches = []
        for m in range(num_matches):
            m1 = m * 2 + 1
            m2 = m * 2 + 2

            p1_empty = (prev_round_name, m1) in empty_matches
            p2_empty = (prev_round_name, m2) in empty_matches

            if p1_empty and p2_empty:
                # Both feeders are empty — this match is also empty
                empty_matches.add((rnd_name, m + 1))
                continue
            elif p1_empty:
                p1_label = "Bye"
                p2_label = bye_winners.get((prev_round_name, m2),
                                           f"Winner {prev_abbrev}-M{m2}")
                match = {
                    "match": m + 1,
                    "player1": p1_label,
                    "player2": p2_label,
                    "notes": f"{p2_label} auto-advances",
                }
                bye_winners[(rnd_name, m + 1)] = p2_label
            elif p2_empty:
                p1_label = bye_winners.get((prev_round_name, m1),
                                           f"Winner {prev_abbrev}-M{m1}")
                p2_label = "Bye"
                match = {
                    "match": m + 1,
                    "player1": p1_label,
                    "player2": p2_label,
                    "notes": f"{p1_label} auto-advances",
                }
                bye_winners[(rnd_name, m + 1)] = p1_label
            else:
                # Resolve feeder labels: use actual name if bye winner, else placeholder
                p1_label = bye_winners.get((prev_round_name, m1),
                                           f"Winner {prev_abbrev}-M{m1}")
                p2_label = bye_winners.get((prev_round_name, m2),
                                           f"Winner {prev_abbrev}-M{m2}")
                match = {
                    "match": m + 1,
                    "player1": p1_label,
                    "player2": p2_label,
                }

            matches.append(match)

        rounds.append({"name": rnd_name, "matches": matches})
        prev_round_name = rnd_name
        prev_match_count = num_matches

    return rounds


def build_playoff_bracket(rows, is_doubles, layout=None):
    """Build bracket structure for a playoff sheet (positions may be empty — just slots)."""
    round_names = get_round_headers(rows, layout)
    # Get draw positions
    data_rows = [r for r in rows if r["_row"] >= 5]
    positions = []
    for r in data_rows:
        a_val = r.get("A")
        if a_val and a_val.isdigit():
            positions.append(int(a_val))
    draw_size = max(positions) if positions else 0

    if draw_size == 0:
        return None

    # For playoffs, positions are typically empty (filled from group results)
    # Build structural bracket with slot labels
    if not round_names:
        return None

    # First round: pair adjacent slots
    first_round_matches = []
    for i in range(1, draw_size + 1, 2):
        first_round_matches.append({
            "match": len(first_round_matches) + 1,
            "player1": f"Slot {i}",
            "player2": f"Slot {i + 1}",
        })

    rounds = [{"name": round_names[0], "matches": first_round_matches}]

    # Later rounds
    abbrev_map = {
        "Round 1": "R1", "Round 2": "R2",
        "Quarter-Final": "QF", "Semi-Final": "SF", "Final": "F",
    }
    prev_name = round_names[0]
    prev_count = len(first_round_matches)

    for rnd_idx in range(1, len(round_names)):
        rnd_name = round_names[rnd_idx]
        prev_abbrev = abbrev_map.get(prev_name, prev_name[:2])
        num_matches = prev_count // 2
        if num_matches < 1:
            break
        matches = []
        for m in range(num_matches):
            matches.append({
                "match": m + 1,
                "player1": f"Winner {prev_abbrev}-M{m * 2 + 1}",
                "player2": f"Winner {prev_abbrev}-M{m * 2 + 2}",
            })
        rounds.append({"name": rnd_name, "matches": matches})
        prev_name = rnd_name
        prev_count = num_matches

    return {
        "format": "elimination",
        "drawSize": draw_size,
        "rounds": rounds,
    }


def extract_roundrobin_schedule(rows, num_players, layout=None):
    """Extract round assignments from the round-robin cross-table matrix.

    Returns dict mapping (pos_i, pos_j) -> round_number (1-based),
    where pos_i < pos_j. Returns empty dict if no round markers found.
    """
    if layout is None:
        layout = detect_layout(rows)
    pcol = layout["player_col"]
    # Matrix columns start right after the player column
    matrix_start = chr(ord(pcol) + 1)

    schedule = {}
    for r in rows:
        a_val = r.get("A", "")
        if not a_val.isdigit():
            continue
        pos_i = int(a_val)
        if pos_i < 1 or pos_i > num_players:
            continue

        for offset in range(num_players):
            col = chr(ord(matrix_start) + offset)
            pos_j = offset + 1
            if pos_j == pos_i:
                continue
            cell = r.get(col, "")
            m = re.match(r"^R(\d+)", cell.strip())
            if m:
                round_num = int(m.group(1))
                key = (min(pos_i, pos_j), max(pos_i, pos_j))
                if key not in schedule:
                    schedule[key] = round_num

    return schedule


def generate_roundrobin_matches(players, is_doubles, round_schedule=None):
    """Generate all-vs-all match pairings for round-robin.

    If round_schedule is provided (dict mapping (pos_i, pos_j) -> round_number),
    each match gets a 'pool_round' field from the schedule.
    """
    matches = []
    match_num = 1
    for i in range(len(players)):
        for j in range(i + 1, len(players)):
            p1 = players[i]
            p2 = players[j]
            name1 = player_label(p1, is_doubles)
            name2 = player_label(p2, is_doubles)
            match = {
                "match": match_num,
                "player1": name1,
                "player2": name2,
            }
            if round_schedule:
                key = (min(p1["position"], p2["position"]),
                       max(p1["position"], p2["position"]))
                pool_round = round_schedule.get(key)
                if pool_round is not None:
                    match["pool_round"] = pool_round
            matches.append(match)
            match_num += 1
    return matches


# ── Club collection helper ───────────────────────────────────────

def collect_clubs(players, is_doubles):
    """Collect unique club names from a player list."""
    clubs = set()
    for p in players:
        if is_doubles:
            for partner in p.get("players", []):
                if partner.get("club"):
                    clubs.add(partner["club"])
        else:
            if p.get("club"):
                clubs.add(p["club"])
    return clubs


# ── Main processing ──────────────────────────────────────────────

def process_workbook(filepath, config):
    wb = openpyxl.load_workbook(filepath, data_only=True)

    output_dir = config["paths"]["divisions_dir"]
    tournament_name = get_tournament_name(config)
    event_names = get_event_names(config)
    level_categories = get_level_categories(config)
    doubles_events = get_doubles_events(config)
    format_overrides = get_format_overrides(config)

    os.makedirs(output_dir, exist_ok=True)

    # First pass: process all sheets and collect data
    sheet_data = {}
    all_clubs = set()

    for sheet_name in wb.sheetnames:
        info = parse_sheet_name(sheet_name, event_names, level_categories, doubles_events)
        if not info:
            continue

        ws = wb[sheet_name]
        rows = read_sheet_rows(ws)
        layout = detect_layout(rows)
        fmt = detect_format(rows, layout)

        # Apply format overrides from config
        div_code = info["code"]
        if div_code in format_overrides:
            fmt = format_overrides[div_code]

        sheet_data[sheet_name] = {
            "info": info,
            "rows": rows,
            "format": fmt,
            "layout": layout,
        }

    # Second pass: build division JSON and link playoffs to main draws
    index_entries = []
    playoff_data = {}  # key: "EVENT LEVEL" -> playoff bracket dict

    # Process playoff sheets first
    for sheet_name, sd in sheet_data.items():
        info = sd["info"]
        if info["draw_type"] != "playoff":
            continue

        rows = sd["rows"]
        bracket = build_playoff_bracket(rows, info["is_doubles"], layout=sd["layout"])
        if bracket:
            playoff_data[info["code"]] = bracket

        # Write standalone playoff JSON
        filename = sheet_to_filename(sheet_name)
        division_json = {
            "tournament": tournament_name,
            "name": f"{info['full_name']} Playoff",
            "code": info["code"],
            "category": info["category"],
            "sheet": sheet_name,
            "draw_type": "playoff",
            "format": "elimination",
            "linked_main_draw": sheet_to_filename(
                sheet_name.replace("-Playoff", "-Main Draw")
            ),
        }
        if bracket:
            division_json["drawSize"] = bracket["drawSize"]
            division_json["rounds"] = bracket["rounds"]
        else:
            division_json["drawSize"] = 0
            division_json["rounds"] = []

        outpath = os.path.join(output_dir, filename)
        with open(outpath, "w", encoding="utf-8") as f:
            json.dump(division_json, f, indent=2, ensure_ascii=False)

        index_entries.append({
            "file": filename,
            "name": division_json["name"],
            "code": info["code"],
            "category": info["category"],
            "draw_type": "playoff",
            "format": "elimination",
        })

    # Process main draw sheets
    for sheet_name, sd in sheet_data.items():
        info = sd["info"]
        if info["draw_type"] != "main_draw":
            continue

        rows = sd["rows"]
        fmt = sd["format"]
        filename = sheet_to_filename(sheet_name)
        is_doubles = info["is_doubles"]

        division_json = {
            "tournament": tournament_name,
            "name": info["full_name"],
            "code": info["code"],
            "category": info["category"],
            "sheet": sheet_name,
            "draw_type": "main_draw",
            "format": fmt,
        }

        div_clubs = set()

        layout = sd["layout"]

        if fmt == "elimination":
            players, draw_size = extract_elimination_players(rows, is_doubles, layout)
            round_names = get_round_headers(rows, layout)
            rounds = build_full_bracket(players, draw_size, round_names, is_doubles)

            division_json["drawSize"] = draw_size
            division_json["players"] = players
            division_json["rounds"] = rounds
            div_clubs = collect_clubs(players, is_doubles)

        elif fmt == "round_robin":
            players = extract_roundrobin_players(rows, is_doubles, layout)
            round_schedule = extract_roundrobin_schedule(
                rows, len(players), layout
            )
            matches = generate_roundrobin_matches(
                players, is_doubles, round_schedule
            )

            division_json["players"] = players
            division_json["matches"] = matches
            div_clubs = collect_clubs(players, is_doubles)

        elif fmt == "group_playoff":
            groups = extract_group_playoff(rows, is_doubles, layout)
            for g in groups:
                round_schedule = g.pop("_round_schedule", None)
                g["matches"] = generate_roundrobin_matches(
                    g["players"], is_doubles, round_schedule
                )
                div_clubs |= collect_clubs(g["players"], is_doubles)

            division_json["groups"] = groups

            # Link playoff bracket if exists
            if info["code"] in playoff_data:
                division_json["playoff"] = playoff_data[info["code"]]
                playoff_file = sheet_to_filename(
                    sheet_name.replace("-Main Draw", "-Playoff")
                )
                division_json["playoff_file"] = playoff_file

        division_json["clubs"] = sorted(div_clubs)
        all_clubs |= div_clubs

        outpath = os.path.join(output_dir, filename)
        with open(outpath, "w", encoding="utf-8") as f:
            json.dump(division_json, f, indent=2, ensure_ascii=False)

        index_entries.append({
            "file": filename,
            "name": info["full_name"],
            "code": info["code"],
            "category": info["category"],
            "draw_type": "main_draw",
            "format": fmt,
        })

    # Write tournament index
    cat_order = get_category_order(config)
    # Sort index by category order, then by code
    def sort_key(e):
        cat_idx = cat_order.index(e["category"]) if e["category"] in cat_order else 99
        return (cat_idx, e["code"], 0 if e["draw_type"] == "main_draw" else 1)

    index_entries.sort(key=sort_key)

    index_json = {
        "tournament": tournament_name,
        "total_divisions": len(index_entries),
        "clubs": sorted(all_clubs),
        "divisions": index_entries,
    }

    index_path = os.path.join(output_dir, "tournament_index.json")
    with open(index_path, "w", encoding="utf-8") as f:
        json.dump(index_json, f, indent=2, ensure_ascii=False)

    return index_json, len(index_entries)


def main(config=None, filepath=None):
    if config is None:
        parser = argparse.ArgumentParser(description="Parse tournament Excel file")
        parser.add_argument("--tournament", default=".",
                            help="Path to tournament directory (default: current dir)")
        parser.add_argument("--file", default=None,
                            help="Path to Excel file (overrides config)")
        args = parser.parse_args()
        config = load_config(args.tournament)
        if args.file:
            filepath = args.file

    if filepath is None:
        # Look for Excel files in the input directory
        input_dir = config["paths"]["input_dir"]
        if os.path.isdir(input_dir):
            xlsx_files = [f for f in os.listdir(input_dir) if f.endswith((".xlsx", ".XLSX"))]
            if xlsx_files:
                filepath = os.path.join(input_dir, xlsx_files[0])

    if filepath is None:
        print("Error: No Excel file specified and none found in input directory.")
        return None, 0

    output_dir = config["paths"]["divisions_dir"]
    print(f"Reading: {filepath}")
    print(f"Output:  {output_dir}/\n")

    index, count = process_workbook(filepath, config)

    print(f"Generated {count} JSON files + tournament_index.json")
    print(f"Total clubs: {len(index['clubs'])}\n")

    # Summary by category
    from collections import defaultdict
    by_cat = defaultdict(list)
    for d in index["divisions"]:
        by_cat[d["category"]].append(d)

    cat_order = get_category_order(config)
    for cat in cat_order:
        divs = by_cat.get(cat, [])
        if divs:
            print(f"  {cat}: {len(divs)} files")
            for d in divs:
                tag = f"[{d['format']}]"
                draw_type = " (playoff)" if d["draw_type"] == "playoff" else ""
                print(f"    {d['file']:40s} {tag:20s}{draw_type}")

    print(f"\nAll files written to: {output_dir}/")

    return index, count


if __name__ == "__main__":
    main()

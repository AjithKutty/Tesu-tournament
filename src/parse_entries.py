"""
Parse an entries-only Excel file and generate random draws.

This is used when only a player list is available (no bracket draws yet).
Reads the entries Excel, determines the appropriate draw format based on
entry count, generates randomized draws, and outputs the same JSON format
as parse_tournament.py / parse_web.py.

Usage:
    python src/parse_entries.py --tournament tournaments/kumpoo-2026
    python src/parse_entries.py --tournament tournaments/kumpoo-2026 --seed 42

Format rules:
    <=6 entries  -> round_robin
    7-12 entries -> group_playoff (groups of 3-4, then elimination playoff)
    >12 entries  -> elimination (bracket padded to next power of 2 with byes)
"""

import argparse
import json
import math
import os
import random
import re
import sys
from itertools import combinations

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config import (load_config, get_tournament_name, get_event_names,
                    get_level_categories, get_doubles_events, get_draw_format)


# ── Excel parsing ────────────────────────────────────────────────

def read_entries(excel_path):
    """Read all entries from the entries Excel file.

    Returns list of dicts: {
        'sheet': str, 'event': str, 'level': str,
        'entries': [{'position': int, 'name': str, 'partner': str|None}]
    }
    """
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    divisions = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Parse sheet name: "EVENT LEVEL - Main Draw"
        m = re.match(r"^(\w+)\s+(\w+)\s*-\s*Main Draw", sheet_name)
        if not m:
            print(f"  Skipping sheet: {sheet_name} (doesn't match pattern)")
            continue

        event = m.group(1)
        level = m.group(2)

        # Find header row (row with "No." in column A)
        header_row = None
        for r in range(1, min(10, ws.max_row + 1)):
            val = ws.cell(r, 1).value
            if val and str(val).strip() == "No.":
                header_row = r
                break

        if header_row is None:
            print(f"  Skipping sheet: {sheet_name} (no header found)")
            continue

        # Read entries
        entries = []
        r = header_row + 1
        while r <= ws.max_row:
            a_val = ws.cell(r, 1).value
            b_val = ws.cell(r, 2).value

            if a_val is not None and str(a_val).strip().isdigit():
                position = int(str(a_val).strip())
                name = str(b_val).strip() if b_val else ""

                # Check if next row is a doubles partner (no number in A, name in B)
                partner = None
                if r + 1 <= ws.max_row:
                    next_a = ws.cell(r + 1, 1).value
                    next_b = ws.cell(r + 1, 2).value
                    if (next_a is None or str(next_a).strip() == "") and next_b and str(next_b).strip():
                        partner = str(next_b).strip()
                        r += 1  # Skip the partner row

                if name:
                    entries.append({
                        'position': position,
                        'name': name,
                        'partner': partner,
                    })
            r += 1

        if entries:
            divisions.append({
                'sheet': sheet_name,
                'event': event,
                'level': level,
                'entries': entries,
            })

    wb.close()
    return divisions


# ── Draw format selection ────────────────────────────────────────
# Format is determined by get_draw_format() from config.py, which checks:
#   1. Per-division override in scheduling.draw_formats.divisions
#   2. Per-category default in scheduling.draw_formats.categories
#   3. Global default in scheduling.draw_formats.default
#   4. Fallback: round_robin if <=6, elimination otherwise


# ── Draw generators ──────────────────────────────────────────────

def make_player_str(entry, is_doubles):
    """Format player name string for match entries."""
    if is_doubles and entry['partner']:
        return f"{entry['name']} / {entry['partner']}"
    return entry['name']


def generate_round_robin(entries, is_doubles):
    """Generate round-robin matches for all entries playing each other."""
    players = [make_player_str(e, is_doubles) for e in entries]
    matches = []
    match_num = 1
    for i, j in combinations(range(len(players)), 2):
        matches.append({
            'match': match_num,
            'player1': players[i],
            'player2': players[j],
        })
        match_num += 1
    return matches


def generate_elimination(entries, is_doubles):
    """Generate an elimination bracket with byes padded to next power of 2.

    Byes are distributed so that the top-ranked players receive them first,
    placed opposite each real player from the bottom of the draw upward.
    This avoids Bye-vs-Bye matches.
    """
    players = [make_player_str(e, is_doubles) for e in entries]
    n = len(players)
    draw_size = 1
    while draw_size < n:
        draw_size *= 2

    num_byes = draw_size - n
    # Place real players in slots, distribute byes so they pair with top players.
    # Standard approach: first `num_byes` players get a bye opponent.
    # Build the draw as pairs: (player, opponent) for each R1 match.
    # Players 0..num_byes-1 face a Bye, players num_byes..n-1 face each other.
    draw = []
    for i in range(draw_size // 2):
        slot_top = i * 2
        slot_bot = i * 2 + 1
        if i < num_byes:
            # This match has a real player vs Bye
            draw.append(players[i])
            draw.append("Bye")
        else:
            # Both slots are real players
            real_idx_top = i
            real_idx_bot = num_byes + (i - num_byes) + (draw_size // 2 - num_byes) + (i - num_byes)
            # Simpler: fill remaining real players in order
            pass

    # Cleaner approach: build the full draw list directly
    draw = [None] * draw_size
    # Place first num_byes players in even positions (top of each pair), byes in odd
    for i in range(num_byes):
        draw[i * 2] = players[i]
        draw[i * 2 + 1] = "Bye"
    # Fill remaining positions with the rest of the players
    remaining = players[num_byes:]
    pos = num_byes * 2
    for p in remaining:
        draw[pos] = p
        pos += 1
    players = draw

    num_rounds = int(math.log2(draw_size))
    rounds = []

    # Round 1
    r1_matches = []
    for i in range(0, draw_size, 2):
        match_num = (i // 2) + 1
        p1 = players[i]
        p2 = players[i + 1]
        m = {'match': match_num, 'player1': p1, 'player2': p2}
        if p1 == "Bye" or p2 == "Bye":
            winner = p2 if p1 == "Bye" else p1
            m['notes'] = f"{winner} auto-advances"
        r1_matches.append(m)

    round_name = _round_name(1, num_rounds)
    rounds.append({'name': round_name, 'matches': r1_matches})

    # Later rounds
    abbrev_map = {}
    for rnd_idx in range(2, num_rounds + 1):
        prev_round_name = _round_name(rnd_idx - 1, num_rounds)
        prev_abbrev = _round_abbrev(prev_round_name)
        current_round_name = _round_name(rnd_idx, num_rounds)
        num_matches = draw_size // (2 ** rnd_idx)

        rnd_matches = []
        for i in range(num_matches):
            match_num = i + 1
            feeder1 = i * 2 + 1
            feeder2 = i * 2 + 2
            rnd_matches.append({
                'match': match_num,
                'player1': f"Winner {prev_abbrev}-M{feeder1}",
                'player2': f"Winner {prev_abbrev}-M{feeder2}",
            })
        rounds.append({'name': current_round_name, 'matches': rnd_matches})

    return rounds, draw_size


def _round_name(round_num, total_rounds):
    """Get round name from 1-based round number and total rounds."""
    remaining = total_rounds - round_num + 1
    if remaining == 1:
        return "Final"
    elif remaining == 2:
        return "Semi-Final"
    elif remaining == 3:
        return "Quarter-Final"
    else:
        return f"Round {round_num}"


def _round_abbrev(round_name):
    """Get round abbreviation for Winner references."""
    abbrevs = {
        "Final": "F",
        "Semi-Final": "SF",
        "Quarter-Final": "QF",
    }
    if round_name in abbrevs:
        return abbrevs[round_name]
    m = re.match(r"Round (\d+)", round_name)
    if m:
        return f"R{m.group(1)}"
    return round_name


def generate_group_playoff(entries, is_doubles, cfg_groups=None, cfg_advancers=None):
    """Generate groups with round-robin, then elimination playoff.

    Args:
        entries: list of player entries
        is_doubles: whether this is a doubles event
        cfg_groups: number of groups (from config), or None to auto-detect
        cfg_advancers: advancers per group (from config), or None to auto-detect
    """
    n = len(entries)

    if cfg_groups is not None:
        num_groups = cfg_groups
    else:
        # Auto-detect: aim for groups of 3-4
        if n <= 8:
            num_groups = 2
        elif n <= 12:
            num_groups = 3 if n <= 9 else 4
        else:
            num_groups = min(n // 3, 6)

    # Distribute players into groups as evenly as possible
    groups = [[] for _ in range(num_groups)]
    for i, entry in enumerate(entries):
        groups[i % num_groups].append(entry)

    group_data = []
    for gi, group_entries in enumerate(groups):
        group_name = f"Group {chr(65 + gi)}"
        players = []
        for pi, entry in enumerate(group_entries):
            p = {
                'position': pi + 1,
                'name': make_player_str(entry, is_doubles),
                'club': None,
                'seed': None,
                'status': None,
            }
            players.append(p)

        matches = generate_round_robin(group_entries, is_doubles)
        group_data.append({
            'name': group_name,
            'players': players,
            'matches': matches,
        })

    # Playoff: top players from each group advance
    if cfg_advancers is not None:
        advancers_per_group = cfg_advancers
    else:
        advancers_per_group = 2 if num_groups <= 4 else 1
    playoff_size = num_groups * advancers_per_group
    bracket_size = 1
    while bracket_size < playoff_size:
        bracket_size *= 2

    num_rounds = int(math.log2(bracket_size))
    playoff_rounds = []

    # First round: Slot placeholders with byes distributed to avoid Bye-vs-Bye
    num_byes = bracket_size - playoff_size
    # Build draw: first num_byes slots get (Slot X, Bye), rest get (Slot X, Slot Y)
    draw_slots = [None] * bracket_size
    slot_counter = 1
    for i in range(num_byes):
        draw_slots[i * 2] = f"Slot {slot_counter}"
        slot_counter += 1
        draw_slots[i * 2 + 1] = "Bye"
    for i in range(num_byes * 2, bracket_size):
        draw_slots[i] = f"Slot {slot_counter}"
        slot_counter += 1

    r1_matches = []
    for i in range(bracket_size // 2):
        match_num = i + 1
        p1 = draw_slots[i * 2]
        p2 = draw_slots[i * 2 + 1]
        m = {'match': match_num, 'player1': p1, 'player2': p2}
        if p1 == "Bye" or p2 == "Bye":
            winner = p2 if p1 == "Bye" else p1
            m['notes'] = f"{winner} auto-advances"
        r1_matches.append(m)

    first_round_name = _round_name(1, num_rounds)
    playoff_rounds.append({'name': first_round_name, 'matches': r1_matches})

    # Later rounds
    for rnd_idx in range(2, num_rounds + 1):
        prev_name = _round_name(rnd_idx - 1, num_rounds)
        prev_abbrev = _round_abbrev(prev_name)
        current_name = _round_name(rnd_idx, num_rounds)
        num_matches = bracket_size // (2 ** rnd_idx)

        rnd_matches = []
        for i in range(num_matches):
            match_num = i + 1
            feeder1 = i * 2 + 1
            feeder2 = i * 2 + 2
            rnd_matches.append({
                'match': match_num,
                'player1': f"Winner {prev_abbrev}-M{feeder1}",
                'player2': f"Winner {prev_abbrev}-M{feeder2}",
            })
        playoff_rounds.append({'name': current_name, 'matches': rnd_matches})

    playoff = {
        'format': 'elimination',
        'drawSize': bracket_size,
        'rounds': playoff_rounds,
    }

    return group_data, playoff


# ── Main ─────────────────────────────────────────────────────────

def main(config=None, seed=None):
    if config is None:
        parser = argparse.ArgumentParser(description="Parse entries and generate random draws")
        parser.add_argument("--tournament", required=True,
                            help="Path to tournament directory")
        parser.add_argument("--seed", type=int, default=None,
                            help="Random seed for reproducible draws")
        args = parser.parse_args()
        config = load_config(args.tournament)
        seed = args.seed

    if seed is not None:
        random.seed(seed)

    tournament_name = get_tournament_name(config)
    event_names = get_event_names(config)
    level_categories = get_level_categories(config)
    doubles_events = get_doubles_events(config)

    input_dir = config["paths"]["input_dir"]
    divisions_dir = config["paths"]["divisions_dir"]
    os.makedirs(divisions_dir, exist_ok=True)

    # Find Excel file
    excel_file = config["tournament"].get("input", {}).get("excel_file")
    if not excel_file:
        print("ERROR: No excel_file configured in tournament.yaml")
        sys.exit(1)

    excel_path = os.path.join(input_dir, excel_file)
    if not os.path.exists(excel_path):
        print(f"ERROR: Excel file not found: {excel_path}")
        sys.exit(1)

    print(f"Reading entries: {excel_path}")
    raw_divisions = read_entries(excel_path)
    print(f"Found {len(raw_divisions)} divisions\n")

    index_entries = []
    all_clubs = set()
    category_files = {}

    for div_data in raw_divisions:
        event = div_data['event']
        level = div_data['level']
        entries = div_data['entries']
        is_doubles = event in doubles_events

        div_code = f"{event} {level}"
        full_name_event = event_names.get(event, event)
        category = level_categories.get(level, level)
        div_name = f"{full_name_event} {level}"

        # Shuffle entries for random draw
        random.shuffle(entries)

        n = len(entries)
        fmt, fmt_params = get_draw_format(config, div_code, category, n)

        params_str = ""
        if fmt_params:
            parts = [f"{k}={v}" for k, v in fmt_params.items()]
            params_str = f" ({', '.join(parts)})"
        print(f"  {div_code:12s} {n:3d} entries -> {fmt}{params_str}")

        # Build player list
        players = []
        for i, entry in enumerate(entries):
            p = {
                'position': i + 1,
                'name': make_player_str(entry, is_doubles),
                'club': None,
                'seed': None,
                'status': None,
            }
            players.append(p)

        # Build division JSON
        div_json = {
            'tournament': tournament_name,
            'name': div_name,
            'code': div_code,
            'category': category,
            'sheet': div_data['sheet'],
            'draw_type': 'main_draw',
            'format': fmt,
            'players': players,
        }

        if fmt == "round_robin":
            div_json['matches'] = generate_round_robin(entries, is_doubles)

        elif fmt == "elimination":
            rounds, draw_size = generate_elimination(entries, is_doubles)
            div_json['drawSize'] = draw_size
            div_json['rounds'] = rounds

        elif fmt == "group_playoff":
            cfg_groups = fmt_params.get('groups')
            cfg_advancers = fmt_params.get('advancers_per_group')
            groups, playoff = generate_group_playoff(
                entries, is_doubles,
                cfg_groups=cfg_groups, cfg_advancers=cfg_advancers,
            )
            div_json['groups'] = groups
            div_json['playoff'] = playoff

            # Also generate the standalone playoff file
            playoff_json = {
                'tournament': tournament_name,
                'name': div_name,
                'code': div_code,
                'category': category,
                'sheet': f"{div_code}-Playoff",
                'draw_type': 'playoff',
                'format': 'elimination',
                'drawSize': playoff['drawSize'],
                'rounds': playoff['rounds'],
                'players': [],
                'clubs': [],
            }
            playoff_filename = f"{div_code.replace(' ', '_')}-Playoff.json"
            playoff_path = os.path.join(divisions_dir, playoff_filename)
            with open(playoff_path, 'w', encoding='utf-8') as f:
                json.dump(playoff_json, f, indent=2, ensure_ascii=False)

            div_json['playoff_file'] = playoff_filename

            index_entries.append({
                'file': playoff_filename,
                'name': div_name,
                'code': div_code,
                'category': category,
                'draw_type': 'playoff',
                'format': 'elimination',
            })

        div_json['clubs'] = []

        # Write division file
        filename = f"{div_code.replace(' ', '_')}-Main_Draw.json"
        filepath = os.path.join(divisions_dir, filename)
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(div_json, f, indent=2, ensure_ascii=False)

        index_entries.append({
            'file': filename,
            'name': div_name,
            'code': div_code,
            'category': category,
            'draw_type': 'main_draw',
            'format': fmt,
        })

        if category not in category_files:
            category_files[category] = []
        category_files[category].append(filename)

    # Sort index: main_draw entries first, grouped by category
    index_entries.sort(key=lambda e: (e['category'], e['code'], e['draw_type']))

    # Write tournament index
    index = {
        'tournament': tournament_name,
        'source': 'entries_excel',
        'divisions': index_entries,
        'clubs': sorted(all_clubs),
    }

    index_path = os.path.join(divisions_dir, "tournament_index.json")
    with open(index_path, 'w', encoding='utf-8') as f:
        json.dump(index, f, indent=2, ensure_ascii=False)

    print(f"\nGenerated {len(index_entries)} JSON files + tournament_index.json")
    print(f"Output: {divisions_dir}/")

    # Summary by category
    for cat in sorted(category_files):
        files = category_files[cat]
        print(f"\n  {cat}: {len(files)} files")
        for fn in sorted(files):
            print(f"    {fn}")


if __name__ == "__main__":
    main()
